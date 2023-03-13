import openpyxl
def excel_to_tuple(worksheet, max_column = 0,max_row = 0):
    result_dict = []
    for row in range(2, max_row+1):
        print(row)
        this_dict = []
        for column in range(1,max_column+1):
            #print(worksheet.cell(row=row, column=column).value)
            this_value = worksheet.cell(row=row, column=column).value
            if this_value == None:
                this_value = ""
            this_dict.append(this_value)
        #print(this_dict)
        result_dict.append(this_dict)

    return result_dict
if __name__ =='__main__':
    path = "auto_send_mail/accept/accept_data.xlsx"
    file  = openpyxl.load_workbook(path)
    sheet = file.worksheets[0]
    receiver = excel_to_tuple(sheet,sheet.max_column,sheet.max_row)
    print(receiver)