#encoding:utf-8
import openpyxl
file_path = "notion_api/data/data.xlsx"
file = openpyxl.load_workbook(file_path)
file_sheet = file.worksheets[0]
print(file_sheet)
print(file_sheet.max_row)
print(file_sheet.max_column)
title = []
for i in range(1,file_sheet.max_column+1):
    title.append(file_sheet.cell(row=1, column=i).value)
print(title)
for i in range(2,file_sheet.max_row+1):
    temp = []
    for j in range(1,file_sheet.max_column+1):
        temp.append(file_sheet.cell(row=i, column=j).value)
    print(len(title))
    print(len(temp))
    with open(f"notion_api/data/{temp[1]}.txt",'w',encoding = "utf8") as f:
        f.write(temp[0])
        f.write('\n')
        print(temp[1])
        f.write(temp[1])
        f.write('\n')
        f.write(temp[2])
        f.write('\n')
        f.write(temp[3])
        f.write('\n')
        for i in range(4,file_sheet.max_column):
            f.write("- "+title[i])
            f.write('\n')
            f.write('\n')
            f.write('\t')
            if temp[i]==None:
                temp[i] = '無'
            f.write(temp[i])
            f.write('\n')
            f.write('\n')

